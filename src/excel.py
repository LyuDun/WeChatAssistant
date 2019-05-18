# -*- coding:utf-8 -*- 
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


class excel(object):
    def __init__(self):
        self._excel = "自动回复.xlsx"
        self.wb = load_workbook(self._excel)
        self.sheet1 = self.wb['Sheet1']
        self.rows = self.sheet1.rows
        self.columns = self.sheet1.columns

    def get_line(self):
        row_length = self.sheet1.max_row
        col_length = self.sheet1.max_column
        return row_length, col_length

    def get_keyword_dict(self):
        col_values = []
        for column in self.columns:
            col_values.append([col.value for col in column])
        return({key: value for key, value in zip(col_values[0][1:], col_values[1][1:])})

    def get_group_list(self):
        grop_list = []
        colC = self.sheet1['C']
        for col in colC:
            if col.value is not None:
                grop_list.append(col.value)
        return(grop_list[1:])

    def add_keyword(self, **kwargs):
        sheet1 = self.sheet1
        row_length, col_length = self.get_line()
        wb = self.wb

        if kwargs is not None:
            for key, value in kwargs.items():
                sheet1.cell(row=row_length+1, column=col_length-2).value = key
                sheet1.cell(row=row_length+1, column=col_length-1).value = value
        wb.save(self._excel)
